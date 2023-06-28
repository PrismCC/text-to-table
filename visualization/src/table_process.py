# 从xls文件中读取数据
import openpyxl
from PIL import Image

xlsx_path = r"table\python_read.xlsx"


def test():
    workbook = openpyxl.load_workbook(xlsx_path)
    print(workbook.sheetnames)


def team_process():
    team_num = 7
    workbook = openpyxl.load_workbook(xlsx_path)
    team_sheet = workbook['Team']
    team_name = "Hawks,Magic,Nets,Wizards,Trail Blaze,Grizzlies,Knicks".split(',')

    points_array = []
    for i in range(team_num):
        line = []
        for j in range(team_num):
            v = team_sheet.cell(row=i + 2, column=j + 2).value
            line.append(0 if v is None else v)
        points_array.append(line)
    print(points_array)

    class Data:
        def __init__(self, name, win, points):
            self.name = name
            self.win = win
            self.points = points

    team_list = []
    for i in range(team_num):
        team = Data(team_name[i], 0, 0)
        for j in range(team_num):
            team.points += points_array[i][j]
            team.win += 1 if points_array[i][j] > points_array[j][i] else 0
        team_list.append(team)

    with open(r"output\teams.csv", "w") as out:
        for i in team_list:
            out.write(f"{i.name},{i.win},{i.points},\n")

    img1 = Image.new('RGBA', (team_num * 100, team_num * 100))
    img2 = Image.new('RGBA', (team_num * 100, team_num * 100))
    for i in range(team_num * 100):
        for j in range(team_num * 100):
            y, x = i // 100, j // 100
            red = (points_array[y][x] + points_array[x][y] - 180) * 3 if points_array[y][x] != 0 else 0
            blue = (points_array[y][x] - points_array[x][y]) * 6 if points_array[y][x] != 0 else 0
            green = (points_array[x][y] - points_array[y][x]) * 6 if points_array[y][x] != 0 else 0
            alpha = 64
            img1.putpixel((i, j), (red, 0, 0, alpha))
            img2.putpixel((i, j), (0, green, blue, alpha))
    img1.save(r'output\team_pic1.png')
    img2.save(r'output\team_pic2.png')


def player_process():
    workbook = openpyxl.load_workbook(xlsx_path)
    player_sheet = workbook['Player']
    line_no = 2
    player_dic = {}

    class Data:
        def __init__(self):
            self.records = 0
            self.scores = 0
            self.assists = 0
            self.rebounds = 0
            self.steals = 0
            self.points = 0
            self.name = ""

        def cal_points(self):
            self.points = (self.scores * 3 + self.assists + self.rebounds + self.steals) // self.records \
                          + self.records * 10

    empty_line = False
    while True:
        name = player_sheet.cell(row=line_no, column=1).value
        if name is not None:
            empty_line = False
        elif empty_line is False:
            empty_line = True
            line_no += 1
            continue
        else:
            break

        if not name in player_dic:
            player_dic[name] = Data()
        player_dic[name].records += 1
        score = player_sheet.cell(row=line_no, column=2).value
        assist = player_sheet.cell(row=line_no, column=3).value
        rebound = player_sheet.cell(row=line_no, column=4).value
        steal = player_sheet.cell(row=line_no, column=5).value
        player_dic[name].scores += score if type(score) == type(1) else 0
        player_dic[name].assists += assist if type(assist) == type(1) else 0
        player_dic[name].rebounds += rebound if type(rebound) == type(1) else 0
        player_dic[name].steals += steal if type(steal) == type(1) else 0
        line_no += 1

    player_list = []
    for player, data in player_dic.items():
        data.cal_points()
        data.name = player.strip()
        player_list.append(data)
    player_list.sort(key=lambda x: x.points, reverse=True)

    with open(r"output\players.csv", "w") as out:
        for i in player_list:
            print(i.name + ": " + i.points.__str__())
            out.write(f"{i.name},{i.points},{i.records},{i.scores},{i.assists},{i.rebounds},{i.steals},\n")


def main():
    team_process()


if __name__ == '__main__':
    main()
